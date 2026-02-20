#!/usr/bin/env python3
#==MCP==
# {
#   "description": "Read XLSX data into JSON via openpyxl.",
#   "schema": {
#     "type": "object",
#     "properties": {
#       "input_path": { "type": "string" },
#       "output_dir": { "type": "string" },
#       "output_filename": { "type": "string" },
#       "max_rows": { "type": "number" }
#     },
#     "required": ["input_path"]
#   }
# }
#==/MCP==
"""Read XLSX workbook content and emit JSON summary."""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    """Read workbook sheets and export a bounded JSON snapshot."""
    parser = argparse.ArgumentParser(description="Read XLSX workbook.")
    parser.add_argument("--input-path", required=True, help="Path to .xlsx file.")
    parser.add_argument("--output-dir", default="/shared/artifacts", help="Artifact output directory.")
    parser.add_argument("--output-filename", help="Optional output JSON filename.")
    parser.add_argument("--max-rows", type=int, default=200, help="Maximum rows per sheet.")
    args = parser.parse_args()

    input_path = Path(args.input_path)
    if not input_path.is_file():
        print(json.dumps({"status": "error", "error": f"Input XLSX not found: {input_path}"}))
        return 1

    try:
        workbook = load_workbook(str(input_path), data_only=True)
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"status": "error", "error": str(exc)}))
        return 1

    data = {}
    # Keep sheet extraction bounded to avoid runaway payload size.
    for sheet in workbook.worksheets:
        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if index > max(args.max_rows, 1):
                break
            rows.append(list(row))
        data[sheet.title] = rows

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = args.output_filename or f"{input_path.stem}.json"
    output_path = output_dir / output_name
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps({"status": "ok", "output_file": str(output_path), "sheet_count": len(data)}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
